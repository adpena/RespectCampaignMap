from openpyxl import load_workbook
import xlrd

import json

import os

import csv

"""
    ITERATING THROUGH SCHOOL DISTRICT DATA,
    IDENTIFYING RANGE & DISTRIBUTION OF VALUES FOR ASSIGNING COLOR BUCKETS.
"""

school_districts_url = (
    "/Users/adpena/PycharmProjects/RespectCampaign/district reports/all_districts/all_files"
)

peims_actual_funding_breakdown = "/Users/adpena/PycharmProjects/CharterCost/2020-2021 PEIMS Actual_funding_breakdown.csv"

# create a list of the district folders in the directory

# print(url)

districts = os.listdir(school_districts_url)


join_legend = {
    "higher_education": {
        "AUSTIN COMMUNITY COLLEGE": "https://join.aft.org/form/austin_community_college_aft/local/06249/accaft-local-6249",
        "COLLEGE OF THE MAINLAND": "https://join.aft.org/form/texas_aft_college_of_the_mainland/local/08041/aft-college-of-the-mainland",
        "LONE STAR COLLEGE": "https://join.aft.org/form/aft_lone_star_college/local/04518/aft-lone-star-college-local-4518",
        "ALAMO COMMUNITY COLLEGE DISTRICT": "https://join.aft.org/form/texas-aft-alamo-community-college-district/local/08041/texas-aft-alamo-community-college-district",
        "OTHER": "https://join.aft.org/form/texas-aft/local/08041/texas-aft-at-large-membership-form",
    },
    "k-12": {},
    "charter": "https://www.texasaft.org/charter-school-membership/",
    "retiree": "https://join.aft.org/form/texas-aft-retiree-plus/local/08041/texas-aft-retiree-plus-membership-form",
    "student": "https://www.texasaft.org/student-membership-application/",
}


def pad_district_number(district_number_local):
    district_number_local = str(district_number_local)
    if len(district_number_local) < 6:
        district_number_local = (
            str(0) * (6 - len(district_number_local))
        ) + district_number_local
    return district_number_local


def process_district_name(district_name_local):
    district_name_local_processed = district_name_local.replace(" County)", ")").replace("Consolidated", "Cons")
    return district_name_local_processed

counter = 0

new_json = {}

with open("Districts2020to2021.geojson", "r") as json_file:
    json_raw = json.load(json_file)

    # print(json_raw)

    for k, v in json_raw.items():
        # print(k)
        if k == "features":
            for feature in v:
                # print(feature)

                district_name = feature["properties"]["NAME"].upper()

                district_number = feature["properties"]["DISTRICT_C"]

                new_json[district_number] = {}

                new_json[district_number]["NAME"] = district_name
                new_json[district_number]["NUMBER"] = district_number

                # TODO: delete properties that you don't need and just make the GeoJSON files larger

                for district in districts:
                    if district != ".DS_Store" and district_name == district.replace(" Data Package.xlsx", ""):

                        # print("MATCH!!!")

                        school_districts_data_package_url = (
                            school_districts_url
                            + "/"
                            + district
                        )

                        # print(data_package_url)

                        wb = load_workbook(
                            filename=school_districts_data_package_url, data_only=True
                        )

                        ws = wb.active

                        # (G15 - G4) / G4

                        # G15: (float(ws['D15'].value) * float(1 + ws['AJ15'].value))

                        # G4: (float(ws['D4'].value) * float(1 + ws['AJ4'].value))

                        percent_change_teacher_pay = round(((
                            (float(ws["D15"].value) * float(1 + ws["AJ15"].value))
                            - (float(ws["D4"].value) * float(1 + ws["AJ4"].value))
                        ) / (float(ws["D4"].value) * float(1 + ws["AJ4"].value))) * 100, 2)

                        feature["properties"][
                            "PercentChangeTeacherPay"
                        ] = percent_change_teacher_pay

                        new_json[district_number]["PercentChangeTeacherPay"] = percent_change_teacher_pay

                        # COLUMN I
                        try:
                            percent_change_beginning_teacher = round(((
                                (float(ws["I15"].value) * float(1 + ws["AJ15"].value))
                                - (float(ws["I4"].value) * float(1 + ws["AJ4"].value))
                            ) / (float(ws["I4"].value) * float(1 + ws["AJ4"].value))) * 100, 2)
                        except Exception:
                            percent_change_beginning_teacher = "N/A"

                        feature["properties"]["PercentChangeBeginningTeacherPay"] = percent_change_beginning_teacher
                        new_json[district_number]["PercentChangeBeginningTeacherPay"] = percent_change_beginning_teacher

                        # COLUMN M
                        try:
                            percent_change_campus_admin = round(((
                                    (float(ws["M15"].value) * float(1 + ws["AJ15"].value))
                                    - (float(ws["M4"].value) * float(1 + ws["AJ4"].value))
                                ) / (float(ws["M4"].value) * float(1 + ws["AJ4"].value))) * 100, 2)
                        except Exception:
                            percent_change_campus_admin = "N/A"

                        feature["properties"]["PercentChangeCampusAdminPay"] = percent_change_campus_admin
                        new_json[district_number]["PercentChangeCampusAdminPay"] = percent_change_campus_admin

                        # COLUMN Q
                        try:
                            percent_change_central_admin = round(((
                                    (float(ws["Q15"].value) * float(1 + ws["AJ15"].value))
                                    - (float(ws["Q4"].value) * float(1 + ws["AJ4"].value))
                                ) / (float(ws["Q4"].value) * float(1 + ws["AJ4"].value))) * 100, 2)
                        except Exception:
                            percent_change_central_admin = "N/A"

                        feature["properties"]["PercentChangeCentralAdminPay"] = percent_change_central_admin
                        new_json[district_number]["PercentChangeCentralAdminPay"] = percent_change_central_admin

                        # COLUMN U
                        try:
                            percent_change_support_staff = round(((
                                    (float(ws["U15"].value) * float(1 + ws["AJ15"].value))
                                    - (float(ws["U4"].value) * float(1 + ws["AJ4"].value))
                                ) / (float(ws["U4"].value) * float(1 + ws["AJ4"].value))) * 100, 2)
                        except Exception:
                            percent_change_support_staff = "N/A"

                        feature["properties"]["PercentChangeSupportStaffPay"] = percent_change_support_staff
                        new_json[district_number]["PercentChangeSupportStaffPay"] = percent_change_support_staff

                        # TODO: Enrollment - GET LONGITUDINAL CHANGE

                        def get_enrollment_change(year_delta, district_number_local):
                            enrollment_url = "/Users/adpena/PycharmProjects/CharterCostTracker/enrollment_reports/statewide_reports_consolidated/"

                            url_suffix = " Statewide Enrollment By District.csv"

                            current_year = "2020-2021"

                            current_year_start = current_year.split("-")[0]

                            range_start = int(current_year_start) - int(year_delta)

                            starting_year = f"{range_start}-{range_start + 1}"

                            # TODO: calculate current_year_enrollment

                            current_year_enrollment = 0

                            with open(
                                f"{enrollment_url}{current_year}{url_suffix}", "r"
                            ) as csvfile:
                                reader = csv.reader(csvfile)

                                header = False
                                header_indexes = {}

                                for row in reader:
                                    if header is False:
                                        header = True
                                        for i, header_label in enumerate(row):
                                            header_indexes[header_label] = i
                                    else:
                                        if (
                                            pad_district_number(
                                                row[header_indexes["District Number"]]
                                            )
                                            == district_number_local
                                        ):
                                            if "<" not in str(
                                                row[header_indexes["Enrollment"]]
                                            ):
                                                current_year_enrollment = float(
                                                    row[header_indexes["Enrollment"]]
                                                )
                                            else:
                                                current_year_enrollment = row[
                                                    header_indexes["Enrollment"]
                                                ]

                            # TODO: calculate starting_year_enrollment

                            starting_year_enrollment = 0

                            with open(
                                f"{enrollment_url}{starting_year}{url_suffix}", "r"
                            ) as csvfile:
                                reader = csv.reader(csvfile)

                                header = False
                                header_indexes = {}

                                for row in reader:
                                    if header is False:
                                        header = True
                                        for i, header_label in enumerate(row):
                                            header_indexes[header_label] = i
                                    else:
                                        if (
                                            pad_district_number(
                                                row[header_indexes["District Number"]]
                                            )
                                            == district_number_local
                                        ):
                                            if "<" not in str(
                                                row[header_indexes["Enrollment"]]
                                            ):
                                                starting_year_enrollment = float(
                                                    row[header_indexes["Enrollment"]]
                                                )
                                            else:
                                                starting_year_enrollment = row[
                                                    header_indexes["Enrollment"]
                                                ]

                            if "<" not in str(
                                current_year_enrollment
                            ) and "<" not in str(starting_year_enrollment):
                                difference = (
                                    str(
                                        round(
                                            float(
                                                (
                                                    current_year_enrollment
                                                    - starting_year_enrollment
                                                )
                                                / starting_year_enrollment
                                            )
                                            * 100,
                                            2,
                                        )
                                    )
                                    + "%"
                                )
                            else:
                                difference = f"(({current_year_enrollment} - {starting_year_enrollment}) / ({starting_year_enrollment}))%"

                            return difference, current_year_enrollment

                        enrollment_change, enrollment = get_enrollment_change(5, district_number)

                        feature["properties"]["EnrollmentChange"] = enrollment_change
                        feature["properties"]["Enrollment"] = enrollment

                        new_json[district_number]["EnrollmentChange"] = enrollment_change
                        new_json[district_number]["Enrollment"] = enrollment

                        def get_respect_campaign_victories(district_name_local):
                            web_text_local_list = []
                            with open(
                                "/Users/adpena/PycharmProjects/RespectCampaignMap/Respect Wage Campaigns For Sharing  - Sheet1.csv"
                            ) as csvfile:
                                reader = csv.reader(csvfile)
                                header = False
                                header_indexes = {}

                                for row in reader:
                                    if header is False:
                                        header = True
                                        for i, header_label in enumerate(row):
                                            header_indexes[header_label] = i
                                    else:
                                        if row[header_indexes["Formal ISD Label"]].upper() == district_name_local.replace(" COUNTY)", ")"):
                                            print("MATCH!")
                                            web_text_local = row[header_indexes["Text for Web"]] + "</ul>"

                                            web_text_local = web_text_local.split("???")
                                            for i, text in enumerate(web_text_local):
                                                if i > 0:

                                                    text = text.strip()
                                                    text_list = text.split(" - ")
                                                    text_final = "<br><br>".join(text_list)
                                                    web_text_local[i] = '<li>' + text_final + "</li>"
                                                else:
                                                    web_text_local[i] = text.strip() + "<ul>"

                                            web_text_local_list.append("<div>" + "".join(web_text_local) + "</div>")

                            return "".join(web_text_local_list)

                        victories = get_respect_campaign_victories(feature["properties"]["NAME"].upper())

                        feature["properties"]["Victories"] = victories
                        new_json[district_number]["Victories"] = victories

                        def get_join_taft_isd_link(district_name_local):
                            with open("/Users/adpena/PycharmProjects/RespectCampaignMap/Join Texas AFT_links by school district.csv", "r") as csvfile:
                                reader = csv.reader(csvfile)
                                header = False

                                header_indexes = {}

                                for row in reader:
                                    if header is False:
                                        header = True

                                        for i, header_label in enumerate(row):
                                            header_indexes[header_label] = i

                                    else:
                                        if row[header_indexes["label"]].upper() == district_name_local.replace(" County)", ")").upper():
                                            # print("LINK MATCH!")
                                            return row[header_indexes["link"]]

                                return ""

                        join_taft_link = get_join_taft_isd_link(district_name)

                        feature["properties"]["JoinTAFT"] = join_taft_link
                        new_json[district_number]["JoinTAFT"] = join_taft_link

                        def get_school_district_data_url(district_name_local):
                            with open("/Users/adpena/PycharmProjects/RespectCampaignMap/School district reports links.csv", "r") as csvfile:
                                reader = csv.reader(csvfile)

                                header = False
                                header_indexes = {}

                                for row in reader:
                                    if header is False:
                                        header = True
                                        for i, header_label in enumerate(row):
                                            header_indexes[header_label.replace("\ufeff", "")] = i

                                    else:
                                        if district_name_local.replace(" County)", ")").upper() == row[header_indexes["DocumentName"]].replace(" Data Package.xlsx", ""):
                                            return row[header_indexes["Links"]]

                        raw_data_link = get_school_district_data_url(district_name)

                        feature["properties"]["RawDataLink"] = raw_data_link
                        new_json[district_number]["RawDataLink"] = raw_data_link

                        def get_legislative_districts(district_name_local):

                            legislative_body_legend = {
                                "House": [],
                                "Senate": [],
                            }

                            districts_by_lege_district = [
                                "/Users/adpena/PycharmProjects/CharterCostTracker/2022 School Districts By Lege District_House.csv",
                                "/Users/adpena/PycharmProjects/CharterCostTracker/2022 School Districts By Lege District_Senate.csv"]

                            for file_url in districts_by_lege_district:
                                legislative_body = file_url.split("_")[1].replace(".csv", "")

                                with open(file_url, "r") as csvfile:
                                    reader = csv.reader(csvfile)

                                    header = False
                                    header_indexes = {}

                                    for row in reader:
                                        if header is False:
                                            header = True
                                            for i, header_label in enumerate(row):
                                                header_indexes[header_label] = i
                                        else:
                                            for i, header_label in enumerate(list(header_indexes.keys())):
                                                if "Name" in header_label:
                                                    if len(row) > i:
                                                        if row[i].upper().replace(" COUNTY)", ")").replace("CONSOLIDATED", "CONS") == district_name_local.upper().replace(" COUNTY)", ")").replace("CONSOLIDATED", "CONS"):
                                                            legislative_body_legend[legislative_body].append(
                                                                (row[0], row[i + 1]))

                            return legislative_body_legend

                        legislative_districts = get_legislative_districts(district_name)

                        feature["properties"]["House"] = legislative_districts["House"]
                        feature["properties"]["Senate"] = legislative_districts["Senate"]

                        new_json[district_number]["House"] = legislative_districts["House"]
                        new_json[district_number]["Senate"] = legislative_districts["Senate"]

                        # TODO: pull in revenue loss from charter transfers from statewide dataset

                        def get_cost_of_charters_and_charter_transfers(district_number_local):
                            default = 0.0

                            with open("/Users/adpena/PycharmProjects/CharterCostTracker/2020-2021 Estimated Revenue Loss to Charters_statewide by district.csv", "r") as csvfile:
                                reader = csv.reader(csvfile)

                                header = False
                                header_indexes = {}

                                for row in reader:
                                    if header is False:
                                        header = True

                                        for i, header_label in enumerate(row):
                                            header_indexes[header_label] = i

                                    else:
                                        if pad_district_number(row[header_indexes["CDN"]]) == district_number_local:
                                            try:
                                                return row[header_indexes["Est. Per Student Revenue Loss"]], row[header_indexes["Total Estimated Revenue Loss to Charters"]], row[header_indexes["Charter Transfers Out"]], row[header_indexes["Total Transfers Out"]]
                                            except Exception:
                                                print("CHARTER COST ERROR:", row)
                                                return default, default, default, default

                            return default, default, default, default

                        feature["properties"]["PerStudentCostOfCharters"], feature["properties"]["CostOfCharters"], feature["properties"]["CharterTransfers"], feature["properties"]["TotalTransfers"] = get_cost_of_charters_and_charter_transfers(district_number)
                        new_json[district_number]["PerStudentCostOfCharters"], new_json[district_number]["CostOfCharters"], new_json[district_number]["CharterTransfers"], new_json[district_number]["TotalTransfers"] = get_cost_of_charters_and_charter_transfers(district_number)

                        def get_current_year_attendance(district_number_local):
                            attendance_local = 0.0

                            xls_url = f"/Users/adpena/PycharmProjects/CharterCostTracker/reports/summaries of finances/excel/2020-2021 {district_number_local} SOF.xls"

                            wb = xlrd.open_workbook(xls_url)
                            sh = wb.sheet_by_index(0)

                            for row in range(sh.nrows):
                                row = sh.row_values(row)
                                if row[1].strip() == "Refined Average Daily Attendance (ADA)":
                                    # print("ADA :)")
                                    attendance_local = float(row[13])
                                    return attendance_local

                            return attendance_local

                        attendance = get_current_year_attendance(district_number)

                        feature["properties"]["Attendance"] = attendance
                        new_json[district_number]["Attendance"] = attendance

                        def get_state_and_local_funding(district_number_local):

                            with open(peims_actual_funding_breakdown, "r") as csvfile:
                                reader = csv.reader(csvfile)

                                header = False
                                header_indexes = {}

                                for row in reader:
                                    if header is False:
                                        header = True

                                        for i, header_label in enumerate(row):
                                            header_indexes[header_label] = i

                                    else:
                                        if pad_district_number(row[header_indexes["District Number"]]) == district_number_local:
                                            return int(row[header_indexes["Local"]].replace("$", "").replace(",", "")), int(row[header_indexes["State"]].replace("$", "").replace(",", "")), int(row[header_indexes["Federal"]].replace("$", "").replace(",", "")), int(row[header_indexes["Other Local"]].replace("$", "").replace(",", "")), int(row[header_indexes["Recapture"]].replace("$", "").replace(",", ""))

                        feature["properties"]["LocalFunding"], feature["properties"]["StateFunding"], feature["properties"]["FederalFunding"], feature["properties"]["OtherLocalFunding"], feature["properties"]["RecaptureAmount"] = get_state_and_local_funding(district_number)
                        new_json[district_number]["LocalFunding"], new_json[district_number]["StateFunding"], new_json[district_number]["FederalFunding"], new_json[district_number]["OtherLocalFunding"], new_json[district_number]["RecaptureAmount"] = get_state_and_local_funding(district_number)

                        def get_charter_cost_link(district_number_local):
                            with open("Cost of Charters reports_2020-2021.csv", "r") as csvfile:
                                reader = csv.reader(csvfile)

                                header = False
                                header_indexes = {}

                                for row in reader:
                                    if header is False:
                                        header = True

                                        for i, header_label in enumerate(row):
                                            header_indexes[header_label] = i

                                    else:

                                        document_name = row[header_indexes["\ufeffDocumentName"]]

                                        if district_number_local in document_name:
                                            return row[header_indexes["Link"]]

                        charter_cost_link = get_charter_cost_link(district_number)

                        feature["properties"]["CharterCostLink"] = charter_cost_link
                        new_json[district_number]["CharterCostLink"] = charter_cost_link

    # with open("DistrictsFinal.geojson", "w", encoding="utf-8") as f:
        # json.dump(json_raw, f, ensure_ascii=False, indent=4)

    with open("DistrictsData.geojson", "w", encoding="utf-8") as f:
        json.dump(new_json, f, ensure_ascii=False, indent=4)

    with open("DistrictsData.js", "w", encoding="utf-8") as f:
        f.write("var DistrictsData = ")
        json.dump(new_json, f, ensure_ascii=False, indent=4)
        f.write(";\n")

    """with open("DistrictsFinal.js", "w", encoding="utf-8") as f:
        f.write("var TexasDistrictsFeatureCollection = ")
        json.dump(json_raw, f, ensure_ascii=False, indent=4)
        f.write(";\n")"""
